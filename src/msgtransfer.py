#!/usr/bin/env python3

import os
import can
import datetime
from openpyxl import load_workbook

# Open can
os.system('sudo ip link set can0 type can bitrate 250000')
os.system('sudo ifconfig can0 up')

# Create can object
can0 = can.interface.Bus(channel = 'can0', bustype = 'socketcan_ctypes')# socketcan_native

#msg = can.Message(arbitration_id=0x123, data=[0, 1, 2, 3, 4, 5, 6, 7], extended_id=False)

# Create the txt file
f = open("/home/pi/Documents/obd_analyzer/data/dev1/can_msg_storage.txt", "w")

# Create file object
wb = load_workbook("/home/pi/Documents/obd_analyzer/data/dev1/can_msg_template.xlsx")

# Grab the active worksheet
ws = wb.active

# Store data object
msg = True
i = 2
print("Start receiving message:")
while msg is not None:
    msg = can0.recv(10.0)
    
    f.write("%s\n" % msg)

    if msg is not None:
        ws["A%d" % i].value = i - 1
        ws["B%d" % i].value = datetime.datetime.fromtimestamp(msg.timestamp)
        ws["C%d" % i].value = msg.arbitration_id
        ws["D%d" % i].value = msg.is_extended_id
        ws["E%d" % i].value = msg.data[0]
        ws["F%d" % i].value = msg.data[1] 
        ws["G%d" % i].value = msg.data[2]
        ws["H%d" % i].value = msg.data[3]
        ws["I%d" % i].value = msg.data[4]
        ws["J%d" % i].value = msg.data[5]
        ws["K%d" % i].value = msg.data[6]
        ws["L%d" % i].value = msg.data[7]
        i += 1
    
#    print(msg)

else:
    f.close()

    wb.save("/home/pi/Documents/obd_analyzer/data/dev1/test.xlsx")
    
    print('Timeout occurred, no message.')

# Close can
os.system('sudo ifconfig can0 down')

